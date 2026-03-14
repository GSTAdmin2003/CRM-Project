"""
Excel template generation for company import.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Column definitions
# Required columns are marked with * in the template
COMPANY_HEADERS = [
    ('Company ID', True),      # Required
    ('Company Name', True),    # Required
    ('Brand Name', False),
    ('Phone', False),
    ('Mobile', False),
    ('Email', False),
    ('Industry', False),
    ('Category', False),
]

# Contact columns: Position header (value=name), then Email, Phone, Mobile
# Example: Director | Email | Phone | Mobile | Accountant | Email | Phone
CONTACT_POSITIONS = ['Director', 'Manager', 'Accountant']
CONTACT_DETAIL_FIELDS = ['Email', 'Phone', 'Mobile']


def generate_company_template(num_contact_groups=3):
    """
    Generate a company Excel import template with formatting and validation.

    Args:
        num_contact_groups: Number of contact groups to include (default 3)

    Returns:
        openpyxl.Workbook: The generated workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Company Import"

    # Define styles
    company_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    contact_fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
    required_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    header_font = Font(bold=True, size=11)
    required_font = Font(bold=True, size=11, color='DC2626')
    example_font = Font(italic=True, color='999999')
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    col = 1

    # Write company headers
    for header, is_required in COMPANY_HEADERS:
        display_header = f'{header} *' if is_required else header
        cell = ws.cell(row=1, column=col, value=display_header)
        cell.font = required_font if is_required else header_font
        cell.fill = required_fill if is_required else company_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        col += 1

    # Write contact group headers
    for group_num in range(1, num_contact_groups + 1):
        # Position column (the header is the position, cell value will be contact name)
        position_name = CONTACT_POSITIONS[group_num - 1] if group_num <= len(CONTACT_POSITIONS) else f'Contact {group_num}'
        cell = ws.cell(row=1, column=col, value=position_name)
        cell.font = header_font
        cell.fill = contact_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        col += 1

        # Detail fields (Email, Phone, Mobile)
        for field in CONTACT_DETAIL_FIELDS:
            cell = ws.cell(row=1, column=col, value=field)
            cell.font = header_font
            cell.fill = contact_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            col += 1

    # Example data row (row 2)
    example_data = [
        # Company fields
        'TAX-123456', 'Acme Corporation', 'Acme', '+1234567890',
        '+1234567891', 'info@acme.com', 'Technology', 'Enterprise',
        # Contact 1: Director (name) | Email | Phone | Mobile
        'Jane Doe', 'jane@acme.com', '+0987654321', '+0987654322',
        # Contact 2: Manager (name) | Email | Phone | Mobile
        'Bob Wilson', 'bob@acme.com', '+1111111111', '+1111111112',
        # Contact 3: Accountant (name) | Email | Phone | Mobile (empty)
        '', '', '', '',
    ]

    for i, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=i, value=value)
        cell.font = example_font
        cell.border = thin_border

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Set column widths
    column_widths = {
        'Company ID': 18,
        'Company Name': 22,
        'Brand Name': 15,
        'Phone': 15,
        'Mobile': 15,
        'Email': 22,
        'Industry': 15,
        'Category': 15,
    }

    # Apply column widths
    for col_idx in range(1, col):
        col_letter = get_column_letter(col_idx)
        header_value = ws.cell(row=1, column=col_idx).value or ''

        # Strip * suffix for lookup
        header_clean = header_value.replace(' *', '')

        # Get width from mapping or default
        if header_clean in column_widths:
            width = column_widths[header_clean]
        elif header_clean == 'Email':
            width = 22
        elif header_clean in ['Phone', 'Mobile']:
            width = 15
        elif header_clean in CONTACT_POSITIONS or header_clean.startswith('Contact'):
            width = 18
        else:
            width = 15

        ws.column_dimensions[col_letter].width = width

    # Set row height for header
    ws.row_dimensions[1].height = 25

    _add_meta_sheet(wb, 'company')
    return wb


# Backwards-compatibility alias
generate_import_template = generate_company_template


def _make_thin_border():
    return Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB'),
    )


def _add_meta_sheet(wb, contact_type):
    meta = wb.create_sheet('_meta')
    meta.sheet_state = 'hidden'
    meta['A1'] = 'contact_type'
    meta['B1'] = contact_type


def generate_combined_template(num_contact_groups=3):
    """
    Generate a combined Excel import template supporting both companies and individuals.

    Includes a visible 'Contact Type *' column (company/individual) so users can mix
    both types in a single file.  The parser reads the per-row value of that column
    and acts accordingly.

    Returns:
        openpyxl.Workbook: The generated workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Import"

    type_fill = PatternFill(start_color='EDE9FE', end_color='EDE9FE', fill_type='solid')
    required_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    company_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    contact_fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
    type_font = Font(bold=True, size=11, color='6D28D9')
    required_font = Font(bold=True, size=11, color='DC2626')
    header_font = Font(bold=True, size=11)
    example_font = Font(italic=True, color='999999')
    thin_border = _make_thin_border()

    def write_header(col, value, font, fill):
        cell = ws.cell(row=1, column=col, value=value)
        cell.font = font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    col = 1

    # Contact Type column — drives per-row behaviour
    write_header(col, 'Contact Type *', type_font, type_fill)
    col += 1

    # Company-side required
    write_header(col, 'Company ID *', required_font, required_fill)
    col += 1

    # All optional company/individual fields
    optional_cols = [
        'Company Name', 'Full Name', 'Brand Name',
        'Phone', 'Mobile', 'Email', 'Industry', 'Category',
    ]
    for header in optional_cols:
        write_header(col, header, header_font, company_fill)
        col += 1

    # Representative groups (company rows only; individuals leave these blank)
    for group_num in range(1, num_contact_groups + 1):
        position_name = CONTACT_POSITIONS[group_num - 1] if group_num <= len(CONTACT_POSITIONS) else f'Contact {group_num}'
        write_header(col, position_name, header_font, contact_fill)
        col += 1
        for field in CONTACT_DETAIL_FIELDS:
            write_header(col, field, header_font, contact_fill)
            col += 1

    # Two example rows — one company, one individual
    example_rows = [
        # company row
        ['company', 'TAX-123456', 'Acme Corporation', '', 'Acme',
         '+1234567890', '+1234567891', 'info@acme.com', 'Technology', 'Enterprise',
         'Jane Doe', 'jane@acme.com', '+0987654321', '+0987654322',
         '', '', '', '', '', '', '', ''],
        # individual row
        ['individual', 'IND001', '', 'John Smith', '',
         '', '+995599000000', 'john@example.com', '', '',
         '', '', '', '', '', '', '', '', '', '', '', ''],
    ]

    for row_num, row_data in enumerate(example_rows, 2):
        for c, value in enumerate(row_data, 1):
            if c > col - 1:
                break
            cell = ws.cell(row=row_num, column=c, value=value)
            cell.font = example_font
            cell.border = thin_border

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 25

    # Column widths
    width_map = {
        'Contact Type *': 16, 'Company ID *': 18,
        'Company Name': 22, 'Full Name': 22, 'Brand Name': 15,
        'Phone': 15, 'Mobile': 15, 'Email': 25, 'Industry': 15, 'Category': 15,
    }
    for col_idx in range(1, col):
        header_val = (ws.cell(row=1, column=col_idx).value or '').replace(' *', '')
        width = width_map.get(header_val) or (22 if 'Email' in header_val else 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    _add_meta_sheet(wb, 'company')   # fallback default; per-row column overrides this
    return wb


# Individual template headers: only the 4 core columns, no rep groups
_INDIVIDUAL_HEADERS = [
    ('Company ID', True),   # legal_id — unique identifier
    ('Full Name', True),    # legal_name alias
    ('Email', False),
    ('Mobile', False),
]


def generate_individual_template():
    """
    Generate an individual-contact Excel import template.

    Visible columns: Company ID *, Full Name *, Email, Mobile.
    No representative groups.  contact_type is embedded in hidden _meta sheet.

    Returns:
        openpyxl.Workbook: The generated workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Individual Import"

    required_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    optional_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    required_font = Font(bold=True, size=11, color='1D4ED8')
    header_font = Font(bold=True, size=11)
    example_font = Font(italic=True, color='999999')
    thin_border = _make_thin_border()

    for col, (header, is_required) in enumerate(_INDIVIDUAL_HEADERS, 1):
        display_header = f'{header} *' if is_required else header
        cell = ws.cell(row=1, column=col, value=display_header)
        cell.font = required_font if is_required else header_font
        cell.fill = required_fill if is_required else optional_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Example row
    for col, value in enumerate(['IND001', 'John Smith', 'john@example.com', '+995599000000'], 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = example_font
        cell.border = thin_border

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 25

    for col_idx, width in enumerate([18, 22, 25, 18], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    _add_meta_sheet(wb, 'individual')
    return wb
