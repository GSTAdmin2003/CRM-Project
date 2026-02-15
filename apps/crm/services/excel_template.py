"""
Excel template generation for lead import.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Column definitions
# Required columns are marked with * in the template
COMPANY_HEADERS = [
    ('Company ID', True),      # Required
    ('Company Name', True),    # Required
    ('Brand Name', False),
    ('Company Phone', False),
    ('Company Email', False),
    ('Industry', False),
    ('Category', False),
]

# Lead fields - simplified for IncomingLead import
LEAD_HEADERS = [
    ('Notes', False),  # Will be used as IncomingLead message
]

# Contact columns: Position header (value=name), then Email, Phone, Mobile
# Example: Director | Email | Phone | Mobile | Accountant | Email | Phone
CONTACT_POSITIONS = ['Director', 'Manager', 'Accountant']
CONTACT_DETAIL_FIELDS = ['Email', 'Phone', 'Mobile']


def generate_import_template(num_contact_groups=3):
    """
    Generate an Excel import template with formatting and validation.

    Args:
        num_contact_groups: Number of contact groups to include (default 3)

    Returns:
        openpyxl.Workbook: The generated workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Lead Import"

    # Define styles
    company_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    lead_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    contact_fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
    required_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    header_font = Font(bold=True, size=11)
    required_font = Font(bold=True, size=11, color='DC2626')
    example_font = Font(italic=True, color='999999')
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    col = 1

    # Write company headers
    for header, is_required in COMPANY_HEADERS:
        display_header = f'{header} *' if is_required else header
        cell = ws.cell(row=1, column=col, value=display_header)
        cell.font = required_font if is_required else header_font
        cell.fill = required_fill if is_required else company_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        col += 1

    # Write lead headers
    for header, is_required in LEAD_HEADERS:
        display_header = f'{header} *' if is_required else header
        cell = ws.cell(row=1, column=col, value=display_header)
        cell.font = required_font if is_required else header_font
        cell.fill = required_fill if is_required else lead_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        col += 1

    # Write contact group headers
    # Format: Position | Email | Phone | Mobile (position header contains contact name)
    for group_num in range(1, num_contact_groups + 1):
        # Position column (the header is the position, cell value will be contact name)
        position_name = CONTACT_POSITIONS[group_num - 1] if group_num <= len(CONTACT_POSITIONS) else f'Contact {group_num}'
        cell = ws.cell(row=1, column=col, value=position_name)
        cell.font = header_font
        cell.fill = contact_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        col += 1

        # Detail fields (Email, Phone, Mobile)
        for field in CONTACT_DETAIL_FIELDS:
            cell = ws.cell(row=1, column=col, value=field)
            cell.font = header_font
            cell.fill = contact_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            col += 1

    # Example data row (row 2)
    example_data = [
        # Company fields
        'TAX-123456', 'Acme Corporation', 'Acme', '+1234567890',
        'info@acme.com', 'Technology', 'Enterprise',
        # Lead fields (Notes = message)
        'Met at trade show, interested in our services',
        # Contact 1: Director (name) | Email | Phone | Mobile
        'Jane Doe', 'jane@acme.com', '+0987654321', '+0987654322',
        # Contact 2: Manager (name) | Email | Phone | Mobile
        'Bob Wilson', 'bob@acme.com', '+1111111111', '+1111111112',
        # Contact 3: Accountant (name) | Email | Phone | Mobile (empty)
        '', '', '', '',
    ]

    for i, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=i, value=value)
        cell.font = example_font
        cell.border = thin_border

    # No special data validations needed for IncomingLead import

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Set column widths (keys without the * suffix)
    column_widths = {
        'Company ID': 18,
        'Company Name': 22,
        'Brand Name': 15,
        'Company Phone': 15,
        'Company Email': 22,
        'Industry': 15,
        'Category': 15,
        'Opportunity Title': 27,
        'Contact Name': 18,
        'Email': 22,
        'Phone': 15,
        'Position': 15,
        'Expected Revenue': 16,
        'Probability (%)': 14,
        'Expected Closing': 16,
        'Source': 15,
        'Notes': 30,
    }

    # Apply column widths
    for col_idx in range(1, col):
        col_letter = get_column_letter(col_idx)
        header_value = ws.cell(row=1, column=col_idx).value or ''

        # Strip * suffix for lookup
        header_clean = header_value.replace(' *', '')

        # Get width from mapping or default
        if header_clean in column_widths:
            width = column_widths[header_clean]
        elif header_clean == 'Email':
            width = 22
        elif header_clean in ['Phone', 'Mobile']:
            width = 15
        elif header_clean in CONTACT_POSITIONS or header_clean.startswith('Contact'):
            # Position columns (contain contact names)
            width = 18
        else:
            width = 15

        ws.column_dimensions[col_letter].width = width

    # Set row height for header
    ws.row_dimensions[1].height = 25

    return wb


def get_template_headers(num_contact_groups=3):
    """
    Get the full list of expected headers for validation.

    Returns:
        list: List of header strings
    """
    headers = [h[0] for h in COMPANY_HEADERS] + [h[0] for h in LEAD_HEADERS]
    for group_num in range(1, num_contact_groups + 1):
        # Position header (contains contact name)
        position_name = CONTACT_POSITIONS[group_num - 1] if group_num <= len(CONTACT_POSITIONS) else f'Contact {group_num}'
        headers.append(position_name)
        # Detail fields
        for field in CONTACT_DETAIL_FIELDS:
            headers.append(field)
    return headers
