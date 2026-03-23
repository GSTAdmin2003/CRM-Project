"""
Template views -- existing Django views preserved during DRF transition.

These views render HTML templates. Business logic that was previously inline
has been moved to CallService where appropriate.
"""

from django import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test

_it_admin_required = user_passes_test(
    lambda u: u.has_role('IT Admin') or u.has_role('Owner') or u.is_staff,
    login_url='/settings/',
)
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST

from ..models import Call, CallRecording, SIPSettings
from ..services import CallService


@login_required
@require_GET
def lookup_inbound_call(request):
    """Read-only lead/contact lookup for an inbound call while it is still ringing.

    Called immediately when the browser receives the SIP INVITE so the agent can
    see who is calling before deciding to answer. Does NOT create any DB records.
    """
    from django.db.models import Q
    from apps.contacts.models import Contact
    from apps.crm.models import Lead
    from core.utils import normalize_phone

    from_number = request.GET.get("from_number", "").strip()
    normalized = normalize_phone(from_number) if from_number else ""

    if not from_number:
        return JsonResponse({"success": False, "error": "from_number required"}, status=400)

    # ── 1. Lookup contacts by phone or mobile ──────────────────────────────
    phone_q = Q(phone=from_number) | Q(mobile=from_number)
    if normalized and normalized != from_number:
        phone_q |= Q(phone=normalized) | Q(mobile=normalized)

    contacts_qs = Contact.objects.filter(phone_q).distinct()
    primary_contact = contacts_qs.first()

    # ── 2. Find opportunities for those contacts ───────────────────────────
    leads = []
    if primary_contact:
        if primary_contact.company_id:
            sibling_ids = Contact.objects.filter(
                company_id=primary_contact.company_id
            ).values_list("id", flat=True)
            leads = list(
                Lead.objects.filter(
                    contact_id__in=sibling_ids,
                    status="converted",
                ).order_by("-updated_at")
            )
        if not leads:
            leads = list(
                Lead.objects.filter(
                    contact__in=contacts_qs,
                    status="converted",
                ).order_by("-updated_at")
            )

    # ── 3. Fallback: search Lead.phone directly ────────────────────────────
    if not leads:
        lead_phone_q = Q(phone=from_number)
        if normalized and normalized != from_number:
            lead_phone_q |= Q(phone=normalized)
        leads = list(
            Lead.objects.filter(lead_phone_q, status="converted")
            .order_by("-updated_at")
        )

    # NOTE: We do NOT create a new Lead here — that only happens when the agent
    # actually answers via register_inbound_call.

    return JsonResponse({
        "success": True,
        "contact_name": primary_contact.name if primary_contact else None,
        "lead_url": leads[0].get_absolute_url() if leads else None,
        "all_leads": [
            {"id": l.id, "title": l.title, "url": l.get_absolute_url()}
            for l in leads
        ],
    })


@login_required
@require_POST
def register_inbound_call(request):
    """Register an inbound call when the agent answers it in the browser.

    Looks up the caller number against Contact.phone/mobile and Lead.phone.
    If found, returns the matching lead(s) for automatic redirect.
    If not found, creates a new opportunity and returns it.
    """
    from django.db.models import Q
    from apps.contacts.models import Contact
    from apps.crm.models import Lead
    from core.utils import normalize_phone

    from_number = request.POST.get("from_number", "").strip()
    normalized = normalize_phone(from_number) if from_number else ""

    # ── 1. Lookup contacts by phone or mobile ──────────────────────────────
    phone_q = Q(phone=from_number) | Q(mobile=from_number)
    if normalized and normalized != from_number:
        phone_q |= Q(phone=normalized) | Q(mobile=normalized)

    contacts_qs = Contact.objects.filter(phone_q).distinct()
    primary_contact = contacts_qs.first()

    # ── 2. Find opportunities for those contacts (inc. company colleagues) ─
    leads = []
    if primary_contact:
        if primary_contact.company_id:
            sibling_ids = Contact.objects.filter(
                company_id=primary_contact.company_id
            ).values_list("id", flat=True)
            leads = list(
                Lead.objects.filter(
                    contact_id__in=sibling_ids,
                    status="converted",
                ).order_by("-updated_at")
            )
        if not leads:
            leads = list(
                Lead.objects.filter(
                    contact__in=contacts_qs,
                    status="converted",
                ).order_by("-updated_at")
            )

    # ── 3. Fallback: search Lead.phone directly ────────────────────────────
    if not leads:
        lead_phone_q = Q(phone=from_number)
        if normalized and normalized != from_number:
            lead_phone_q |= Q(phone=normalized)
        leads = list(
            Lead.objects.filter(lead_phone_q, status="converted")
            .order_by("-updated_at")
        )

    # ── 4. Nothing found — create opportunity ─────────────────────────────
    if not leads:
        display_name = primary_contact.name if primary_contact else from_number
        new_lead = Lead.objects.create(
            title=f"Inbound call from {display_name}",
            phone=from_number,
            full_name=display_name,
            contact=primary_contact,
            source="cold_call",
            status="converted",
            assigned_to=request.user,
            created_by=request.user,
        )
        leads = [new_lead]

    primary_lead = leads[0]

    # ── 5. Find or create Call record ──────────────────────────────────────
    # The ARI event handler already creates the Call when Asterisk receives the
    # inbound channel. We look that up and update it rather than creating a
    # duplicate (which would violate the asterisk_channel_id unique constraint).
    import datetime
    import uuid

    cutoff = timezone.now() - datetime.timedelta(minutes=5)
    phone_variants = [from_number]
    if normalized and normalized != from_number:
        phone_variants.append(normalized)

    call = (
        Call.objects.filter(
            direction="inbound",
            from_number__in=phone_variants,
            started_at__gte=cutoff,
            status__in=["ringing", "answered"],
        )
        .order_by("-started_at")
        .first()
    )

    if call:
        call.contact = primary_contact
        call.opportunity = primary_lead
        call.user = request.user
        call.answered_at = timezone.now()
        call.status = "answered"
        call.save(update_fields=["contact", "opportunity", "user", "answered_at", "status", "updated_at"])
    else:
        # Fallback: ARI call not found (e.g. call answered before ARI event arrived)
        call = Call.objects.create(
            asterisk_channel_id=f"inbound-{uuid.uuid4().hex[:16]}",
            direction="inbound",
            from_number=from_number,
            to_number=getattr(request.user, "extension", None) or "100",
            status="answered",
            user=request.user,
            contact=primary_contact,
            opportunity=primary_lead,
            started_at=timezone.now(),
            answered_at=timezone.now(),
        )

    # ── 6. Create Activity linked to the lead ─────────────────────────────
    activity_id = None
    try:
        from apps.activities.models import Activity, ActivityType, PhoneCallExtension
        phone_call_type = ActivityType.objects.filter(name="Phone Call", is_active=True).first()
        if phone_call_type:
            display_name = primary_contact.name if primary_contact else from_number
            activity = Activity.objects.create(
                activity_type=phone_call_type,
                title=f"Inbound call from {display_name}",
                scheduled_date=timezone.now().date(),
                status="planned",
                assigned_to=request.user,
                created_by=request.user,
                lead=primary_lead,
            )
            call.activity = activity
            call.save(update_fields=["activity", "updated_at"])

            PhoneCallExtension.objects.create(
                activity=activity,
                ari_call=call,
                direction="inbound",
                call_status="answered",
                from_number=from_number,
                to_number=getattr(request.user, "extension", None) or "100",
                asterisk_channel_id=call.asterisk_channel_id,
                user=request.user,
                started_at=call.started_at,
                answered_at=call.answered_at,
            )
            activity_id = activity.id
    except Exception:
        pass

    return JsonResponse({
        "success": True,
        "call_id": call.id,
        "activity_id": activity_id,
        "lead_id": primary_lead.id,
        "lead_url": primary_lead.get_absolute_url(),
        "contact_name": primary_contact.name if primary_contact else None,
        "all_leads": [
            {"id": l.id, "title": l.title, "url": l.get_absolute_url()}
            for l in leads
        ],
    })


@login_required
@require_POST
def record_missed_call(request):
    """Record an inbound call that rang but was never answered.

    Called from the browser when the SIP INVITE times out or is rejected by
    the agent clicking Reject.  Creates (or updates) the Call record with
    status='no_answer' and logs a 'Missed call' Activity against the lead.
    """
    import datetime
    import uuid

    from django.db.models import Q

    from apps.contacts.models import Contact
    from apps.crm.models import Lead
    from core.utils import normalize_phone

    from_number = request.POST.get("from_number", "").strip()
    if not from_number:
        return JsonResponse({"success": False, "error": "from_number required"}, status=400)

    normalized = normalize_phone(from_number) if from_number else ""

    # ── 1. Contact lookup ──────────────────────────────────────────────────
    phone_q = Q(phone=from_number) | Q(mobile=from_number)
    if normalized and normalized != from_number:
        phone_q |= Q(phone=normalized) | Q(mobile=normalized)
    contacts_qs = Contact.objects.filter(phone_q).distinct()
    primary_contact = contacts_qs.first()

    # ── 2. Lead lookup ─────────────────────────────────────────────────────
    leads = []
    if primary_contact:
        if primary_contact.company_id:
            sibling_ids = Contact.objects.filter(
                company_id=primary_contact.company_id
            ).values_list("id", flat=True)
            leads = list(
                Lead.objects.filter(
                    contact_id__in=sibling_ids, status="converted"
                ).order_by("-updated_at")
            )
        if not leads:
            leads = list(
                Lead.objects.filter(
                    contact__in=contacts_qs, status="converted"
                ).order_by("-updated_at")
            )
    if not leads:
        lead_phone_q = Q(phone=from_number)
        if normalized and normalized != from_number:
            lead_phone_q |= Q(phone=normalized)
        leads = list(
            Lead.objects.filter(lead_phone_q, status="converted").order_by(
                "-updated_at"
            )
        )

    primary_lead = leads[0] if leads else None

    # ── 3. Update ARI call record (if it exists) or create a new one ───────
    cutoff = timezone.now() - datetime.timedelta(minutes=5)
    phone_variants = [from_number]
    if normalized and normalized != from_number:
        phone_variants.append(normalized)

    call = (
        Call.objects.filter(
            direction="inbound",
            from_number__in=phone_variants,
            started_at__gte=cutoff,
            status="ringing",
        )
        .order_by("-started_at")
        .first()
    )

    if call:
        call.status = "no_answer"
        call.ended_at = timezone.now()
        call.contact = primary_contact
        call.opportunity = primary_lead
        call.save(update_fields=["status", "ended_at", "contact", "opportunity", "updated_at"])
    else:
        call = Call.objects.create(
            asterisk_channel_id=f"missed-{uuid.uuid4().hex[:16]}",
            direction="inbound",
            from_number=from_number,
            status="no_answer",
            contact=primary_contact,
            opportunity=primary_lead,
            started_at=timezone.now(),
            ended_at=timezone.now(),
        )

    # ── 4. Create a 'Missed call' Activity ─────────────────────────────────
    if primary_lead:
        try:
            from apps.activities.models import Activity, ActivityType, PhoneCallExtension

            phone_call_type = ActivityType.objects.filter(
                name="Phone Call", is_active=True
            ).first()
            if phone_call_type:
                display = primary_contact.name if primary_contact else from_number
                activity = Activity.objects.create(
                    activity_type=phone_call_type,
                    title=f"Missed inbound call from {display}",
                    scheduled_date=timezone.now().date(),
                    status="completed",
                    outcome="Missed call — agent did not answer",
                    assigned_to=request.user,
                    created_by=request.user,
                    lead=primary_lead,
                )
                PhoneCallExtension.objects.create(
                    activity=activity,
                    ari_call=call,
                    direction="inbound",
                    call_status="no_answer",
                    from_number=from_number,
                    to_number=getattr(request.user, "extension", None) or "100",
                    asterisk_channel_id=call.asterisk_channel_id,
                    user=request.user,
                    started_at=call.started_at,
                )
        except Exception:
            pass

    return JsonResponse({"success": True})


@login_required
@require_POST
def call_ended(request, pk):
    """Mark an inbound call as ended and trigger recording processing."""
    try:
        call = Call.objects.get(pk=pk, user=request.user)
    except Call.DoesNotExist:
        return JsonResponse({"error": "Call not found"}, status=404)

    if call.status in ("ended", "failed"):
        return JsonResponse({"success": True, "call_id": call.id})

    call.status = "ended"
    call.ended_at = timezone.now()
    if call.answered_at:
        call.duration = int((call.ended_at - call.answered_at).total_seconds())
    call.save()

    # Recording is triggered by the ARI ChannelDestroyed event — not here.

    return JsonResponse({"success": True, "call_id": call.id})


@login_required
@require_POST
def upload_browser_recording(request, pk):
    """Accept a browser-recorded WebM audio file as fallback when Asterisk recording is unavailable."""
    try:
        call = Call.objects.get(pk=pk, user=request.user)
    except Call.DoesNotExist:
        return JsonResponse({"success": False, "error": "Call not found"}, status=404)

    # Skip if Asterisk already saved a recording for this call
    if hasattr(call, "recording") and call.recording.file:
        return JsonResponse({"success": True, "skipped": True})

    audio_file = request.FILES.get("recording")
    if not audio_file:
        return JsonResponse({"success": False, "error": "No recording file"}, status=400)

    recording = CallRecording.objects.create(
        call=call,
        file=audio_file,
        duration=call.duration,
        file_size=audio_file.size,
    )

    # Auto-trigger transcription (mirrors process_recording logic)
    try:
        from ..models import CallTranscript, LANGUAGE_TO_STT_CODE
        from ..tasks import transcribe_call
        from apps.user_settings.models.general import SystemConfiguration

        auto_transcribe = SystemConfiguration.get_setting("transcription_auto_enabled", False)

        lang_code = ""
        _contact = (
            call.contact
            if call.contact_id
            else (
                call.opportunity.contact
                if call.opportunity_id and call.opportunity.contact_id
                else None
            )
        )
        if _contact:
            lang_code = LANGUAGE_TO_STT_CODE.get(_contact.effective_language, "")
        if not lang_code:
            default_lang = SystemConfiguration.get_setting("default_preferred_language", "en")
            lang_code = LANGUAGE_TO_STT_CODE.get(default_lang, "") or "en"

        transcript, _ = CallTranscript.objects.get_or_create(call=call)
        transcript.language_code = lang_code
        transcript.save(update_fields=["language_code", "updated_at"])

        if auto_transcribe:
            task = transcribe_call.delay(call.id)
            transcript.celery_task_id = task.id
            transcript.save(update_fields=["celery_task_id", "updated_at"])
    except Exception:
        pass  # transcription is best-effort

    return JsonResponse({"success": True, "skipped": False})


@login_required
@require_POST
def initiate_call(request):
    """Initiate an outbound call via AJAX"""
    to_number = request.POST.get("to_number", "").strip()
    contact_id = request.POST.get("contact_id")
    opportunity_id = request.POST.get("opportunity_id")

    try:
        call = CallService.initiate_call(
            user=request.user,
            phone_number=to_number,
            contact_id=int(contact_id) if contact_id else None,
            lead_id=int(opportunity_id) if opportunity_id else None,
        )
        return JsonResponse({
            "success": True,
            "call_id": call.id,
            "channel_id": call.asterisk_channel_id,
            "activity_id": call.activity_id,
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400 if "required" in str(e) else 500)


@login_required
@require_POST
def hangup_call(request, pk):
    """Hangup an active call"""
    try:
        CallService.hangup_call(call_pk=pk, user=request.user)
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
@require_POST
def answer_call(request, pk):
    """Answer a ringing call"""
    try:
        CallService.answer_call(call_pk=pk, user=request.user)
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400 if "not ringing" in str(e) else 500)


@login_required
@require_GET
def call_status(request, pk):
    """Get current call status (for polling)"""
    try:
        status_data = CallService.get_call_status(call_pk=pk)
        return JsonResponse(status_data)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=404)


@login_required
@require_POST
def update_call_notes(request, pk):
    """Update call notes"""
    notes = request.POST.get("notes", "")

    try:
        CallService.update_call_notes(call_pk=pk, user=request.user, notes=notes)
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=404)


@login_required
@require_POST
def link_call_to_contact(request, pk):
    """Link a call to a contact"""
    contact_id = request.POST.get("contact_id")

    if contact_id:
        try:
            CallService.link_call_to_contact(call_pk=pk, contact_id=int(contact_id))
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=404)

    return JsonResponse({"success": True})


@login_required
@require_POST
def link_call_to_opportunity(request, pk):
    """Link a call to an opportunity"""
    opportunity_id = request.POST.get("opportunity_id")

    if opportunity_id:
        try:
            CallService.link_call_to_opportunity(call_pk=pk, lead_id=int(opportunity_id))
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=404)

    return JsonResponse({"success": True})


@login_required
def recording_download(request, pk):
    """Download call recording"""
    recording = get_object_or_404(CallRecording, pk=pk)

    response = HttpResponse(recording.file, content_type="audio/mpeg")
    response["Content-Disposition"] = (
        f'attachment; filename="{recording.call.asterisk_uniqueid}.mp3"'
    )
    return response


@login_required
def active_calls(request):
    """Get list of active calls (for dashboard widget)"""
    calls_data = CallService.get_active_calls(user=request.user)
    return JsonResponse({"calls": calls_data})


def _calls_xlsx_response(ids):
    """Build a formatted .xlsx response for the given call IDs."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    calls = Call.objects.filter(id__in=ids).select_related("contact", "user")
    headers = ["Direction", "From", "To", "Contact", "Status", "Duration (s)", "User", "Started At"]
    rows = [
        [
            call.direction,
            call.from_number or "",
            call.to_number or "",
            call.contact.name if call.contact else "",
            call.status,
            call.duration or "",
            call.user.get_full_name() if call.user else "",
            call.started_at.strftime("%Y-%m-%d %H:%M") if call.started_at else "",
        ]
        for call in calls
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Calls"
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    for col_idx, h in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(h)
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="calls_export.xlsx"'
    return response


@login_required
@require_POST
def calls_bulk_action(request):
    """Handle bulk actions on calls."""
    import csv

    from core.exceptions import NotFoundError, PermissionDeniedError, ValidationError

    action = request.POST.get("action")
    ids = [int(i) for i in request.POST.getlist("selected_ids") if i.isdigit()]
    if not ids:
        messages.warning(request, "No items selected.")
        return redirect("activities:dashboard")
    try:
        if action == "export_csv":
            calls = Call.objects.filter(id__in=ids).select_related("contact", "user")
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="calls_export.csv"'
            writer = csv.writer(response)
            writer.writerow(
                ["Direction", "From", "To", "Contact", "Status", "Duration (s)", "User", "Started At"]
            )
            for call in calls:
                writer.writerow([
                    call.direction,
                    call.from_number or "",
                    call.to_number or "",
                    call.contact.name if call.contact else "",
                    call.status,
                    call.duration or "",
                    call.user.get_full_name() if call.user else "",
                    call.started_at.strftime("%Y-%m-%d %H:%M") if call.started_at else "",
                ])
            return response
        elif action == "export_excel":
            return _calls_xlsx_response(ids)
        elif action == "delete":
            count = CallService.bulk_delete(ids=ids, user=request.user)
            messages.success(request, f"{count} call{'s' if count != 1 else ''} deleted.")
        else:
            messages.warning(request, "Unknown action.")
    except (NotFoundError, PermissionDeniedError, ValidationError) as e:
        messages.error(request, str(e))
    return redirect("activities:dashboard")


WEEKDAY_CHOICES = [
    ("mon", "Monday"),
    ("tue", "Tuesday"),
    ("wed", "Wednesday"),
    ("thu", "Thursday"),
    ("fri", "Friday"),
    ("sat", "Saturday"),
    ("sun", "Sunday"),
]


class SIPSettingsForm(forms.ModelForm):
    password = forms.CharField(
        widget=forms.PasswordInput(render_value=True, attrs={"autocomplete": "new-password"}),
        required=False,
    )
    working_days_choices = forms.MultipleChoiceField(
        choices=WEEKDAY_CHOICES,
        widget=forms.CheckboxSelectMultiple,
        required=False,
        label="Working Days",
    )

    class Meta:
        model = SIPSettings
        fields = [
            "server_ip", "server_port", "username", "caller_id", "is_active",
            "hold_music", "welcome_sound", "non_working_hours_sound",
            "working_hours_start", "working_hours_end",
        ]
        widgets = {
            "working_hours_start": forms.TimeInput(
                format="%H:%M",
                attrs={"type": "time", "class": "form-control"},
            ),
            "working_hours_end": forms.TimeInput(
                format="%H:%M",
                attrs={"type": "time", "class": "form-control"},
            ),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance and self.instance.pk:
            self.fields["password"].initial = self.instance.password
            # Populate working_days_choices from CSV field
            if self.instance.working_days:
                self.fields["working_days_choices"].initial = [
                    d.strip() for d in self.instance.working_days.split(",") if d.strip()
                ]
        else:
            self.fields["password"].required = True

    def _validate_sound_field(self, field_name):
        sound = self.cleaned_data.get(field_name)
        if sound and hasattr(sound, "name"):
            ext = sound.name.rsplit(".", 1)[-1].lower() if "." in sound.name else ""
            if ext not in ("mp3", "wav"):
                self.add_error(field_name, "Only mp3 and wav files are supported.")

    def clean(self):
        cleaned = super().clean()

        # Validate sound file extensions
        self._validate_sound_field("hold_music")
        self._validate_sound_field("welcome_sound")
        self._validate_sound_field("non_working_hours_sound")

        # Convert working_days_choices to CSV string
        days = cleaned.get("working_days_choices", [])
        cleaned["working_days"] = ",".join(days)

        # Validate working hours: both start+end required together
        start = cleaned.get("working_hours_start")
        end = cleaned.get("working_hours_end")
        if start and not end:
            self.add_error("working_hours_end", "End time is required when start time is set.")
        if end and not start:
            self.add_error("working_hours_start", "Start time is required when end time is set.")
        if (start and end) and not days:
            self.add_error(
                "working_days_choices",
                "Select at least one working day when hours are set."
            )

        server_ip = cleaned.get("server_ip")
        server_port = cleaned.get("server_port")
        username = cleaned.get("username")
        password = cleaned.get("password")

        # For existing settings with no new password, use stored password
        if not password and self.instance and self.instance.pk:
            password = self.instance.password

        if not all([server_ip, server_port, username, password]):
            return cleaned

        # Only validate if the trunk will be active
        if not cleaned.get("is_active", True):
            return cleaned

        # Skip SIP validation if credentials haven't changed (avoids blocking
        # saves for schedule/sound-only changes when SIP server is slow/flaky)
        if self.instance and self.instance.pk:
            creds_changed = (
                server_ip != self.instance.server_ip
                or server_port != self.instance.server_port
                or username != self.instance.username
                or (cleaned.get("password") and cleaned["password"] != self.instance.password)
                or cleaned.get("is_active") != self.instance.is_active
            )
            if not creds_changed:
                return cleaned

        from ..sip_validator import validate_sip_credentials

        is_valid, error = validate_sip_credentials(
            server_ip, server_port, username, password
        )
        if not is_valid:
            raise forms.ValidationError(f"SIP credential test failed: {error}")

        return cleaned

    def save(self, commit=True):
        instance = super().save(commit=False)
        password = self.cleaned_data.get("password")
        if password:
            instance.password = password  # uses the encrypting setter
        # Explicitly set schedule fields from cleaned_data
        instance.working_days = self.cleaned_data.get("working_days", "")
        instance.working_hours_start = self.cleaned_data.get("working_hours_start")
        instance.working_hours_end = self.cleaned_data.get("working_hours_end")
        if commit:
            instance.save()
        return instance


@login_required
def sip_settings_view(request):
    """SIP credentials settings page"""
    from ..asterisk_config import apply_moh_settings, apply_schedule_settings, apply_sip_settings

    try:
        sip = request.user.sip_settings
    except SIPSettings.DoesNotExist:
        sip = None

    if request.method == "POST":
        if "delete" in request.POST:
            if sip:
                sip.delete()
                apply_sip_settings(None)
                apply_schedule_settings(None)
                messages.success(request, "SIP credentials deleted. Asterisk config cleared.")
            return redirect("settings:voip:sip_credentials")

        # Handle hold music removal
        if "remove_hold_music" in request.POST and sip and sip.hold_music:
            sip.hold_music.delete(save=False)
            sip.hold_music = None
            sip.save()
            apply_moh_settings(sip)
            messages.success(request, "Hold music removed. Default music will be used.")
            return redirect("settings:voip:sip_credentials")

        # Handle welcome sound removal
        if "remove_welcome_sound" in request.POST and sip and sip.welcome_sound:
            sip.welcome_sound.delete(save=False)
            sip.welcome_sound = None
            sip.save()
            apply_schedule_settings(sip)
            messages.success(request, "Welcome sound removed.")
            return redirect("settings:voip:sip_credentials")

        # Handle non-working hours sound removal
        if "remove_non_working_sound" in request.POST and sip and sip.non_working_hours_sound:
            sip.non_working_hours_sound.delete(save=False)
            sip.non_working_hours_sound = None
            sip.save()
            apply_schedule_settings(sip)
            messages.success(request, "Non-working hours sound removed.")
            return redirect("settings:voip:sip_credentials")

        # Snapshot current trunk values BEFORE form binding modifies the instance
        old_trunk = None
        if sip:
            old_trunk = {
                "server_ip": sip.server_ip,
                "server_port": sip.server_port,
                "username": sip.username,
                "is_active": sip.is_active,
                "password": sip.password,
            }

        form = SIPSettingsForm(request.POST, request.FILES, instance=sip)
        if form.is_valid():
            # Detect if SIP trunk credentials changed (to avoid unnecessary PJSIP reload
            # which drops the trunk registration and causes inbound calls to fail briefly)
            trunk_changed = old_trunk is None  # new settings = always apply trunk
            if old_trunk:
                for field in ("server_ip", "server_port", "username", "is_active"):
                    if form.cleaned_data.get(field) != old_trunk[field]:
                        trunk_changed = True
                        break
                # Password change
                pwd = form.cleaned_data.get("password")
                if pwd and pwd != old_trunk["password"]:
                    trunk_changed = True

            obj = form.save(commit=False)
            obj.user = request.user
            obj.save()

            # Apply MOH settings if hold music was uploaded
            if "hold_music" in request.FILES:
                apply_moh_settings(obj)

            # Always apply schedule settings on save — the generated dialplan
            # depends on working hours + sounds, and it's cheap to regenerate.
            # This only reloads pbx_config (dialplan), NOT pjsip, so it's safe.
            apply_schedule_settings(obj)

            # Only reload PJSIP (trunk config) if credentials actually changed.
            # Reloading res_pjsip.so drops the trunk registration temporarily,
            # causing inbound calls to fail until the trunk re-registers.
            if trunk_changed:
                if apply_sip_settings(obj):
                    messages.success(request, "SIP credentials saved and Asterisk reloaded.")
                else:
                    messages.warning(
                        request,
                        "SIP credentials saved but Asterisk reload failed. "
                        "You may need to restart the Asterisk container.",
                    )
            else:
                messages.success(request, "Settings saved.")
            return redirect("settings:voip:sip_credentials")
    else:
        form = SIPSettingsForm(instance=sip)

    import json
    from django.core.serializers.json import DjangoJSONEncoder
    init_data = {
        "apiUrls": {
            "sipSettings": "/settings/api/sip/",
            "voipConfig": "/settings/api/voip/config/",
            "voipSounds": "/settings/api/voip/sounds/",
            "workingHours": "/settings/api/voip/working-hours/",
        }
    }
    return render(request, "calls/sip_settings.html", {
        "form": form,
        "sip_settings": sip,
        "weekday_choices": WEEKDAY_CHOICES,
        "current_section": "voip",
        "current_page": "sip_credentials",
        "init_data_json": json.dumps(init_data, cls=DjangoJSONEncoder),
    })


@login_required
@require_POST
def start_transcription(request, pk):
    """Set language and trigger transcription (or re-transcription) for a call."""
    from ..models import CallTranscript, LANGUAGE_TO_STT_CODE
    from ..tasks import transcribe_call

    call = get_object_or_404(Call, pk=pk)

    language_short = request.POST.get('language', '')
    lang_code = LANGUAGE_TO_STT_CODE.get(language_short, '')
    if not lang_code:
        return JsonResponse({'error': 'Invalid or missing language'}, status=400)

    transcript, _ = CallTranscript.objects.get_or_create(call=call)

    # Block if already processing to avoid double-submission
    if transcript.status == CallTranscript.STATUS_PROCESSING:
        return JsonResponse({'error': 'Transcription already in progress'}, status=400)

    transcript.language_code = lang_code
    transcript.status = CallTranscript.STATUS_PENDING
    transcript.error_message = ''
    transcript.save(update_fields=['language_code', 'status', 'error_message', 'updated_at'])

    result = transcribe_call.delay(call.id)
    transcript.celery_task_id = result.id
    transcript.save(update_fields=['celery_task_id'])

    return JsonResponse({'success': True})


@login_required
@require_POST
def cancel_transcription(request, pk):
    """Cancel a pending or processing transcription."""
    from ..models import CallTranscript

    call = get_object_or_404(Call, pk=pk)

    try:
        transcript = call.transcript
    except CallTranscript.DoesNotExist:
        return JsonResponse({'error': 'No transcript found'}, status=404)

    if transcript.status not in (CallTranscript.STATUS_PENDING, CallTranscript.STATUS_PROCESSING):
        return JsonResponse({'error': 'Transcript is not pending or processing'}, status=400)

    # Revoke the Celery task if we have its ID
    if transcript.celery_task_id:
        try:
            from celery import current_app
            current_app.control.revoke(
                transcript.celery_task_id, terminate=True, signal='SIGTERM'
            )
        except Exception:
            pass  # Best-effort — reset status regardless

    transcript.status = CallTranscript.STATUS_PENDING
    transcript.language_code = ''
    transcript.celery_task_id = ''
    transcript.error_message = ''
    transcript.save(update_fields=['status', 'language_code', 'celery_task_id', 'error_message', 'updated_at'])

    return JsonResponse({'success': True})


@login_required
@require_POST
def start_ai_analysis(request, pk):
    """Trigger (or re-trigger) Claude AI analysis for a completed call transcript."""
    from ..models import CallAnalysis
    from ..tasks import analyze_call_with_claude

    call = get_object_or_404(Call, pk=pk)

    # Must have a completed transcript
    transcript = getattr(call, 'transcript', None)
    if not transcript or transcript.status != 'completed':
        return render(request, 'calls/_analysis_section.html', {
            'call': call,
            'analysis': getattr(call, 'analysis', None),
            'error': 'Transcript must be completed before AI analysis can run.',
        })

    analysis, _ = CallAnalysis.objects.get_or_create(call=call)
    analysis.status = CallAnalysis.STATUS_PENDING
    analysis.error_message = ''
    analysis.analysis_text = ''
    analysis.suggested_activities = []
    analysis.created_activity_ids = []
    analysis.save()

    result = analyze_call_with_claude.delay(call.pk)
    analysis.celery_task_id = result.id
    analysis.save(update_fields=['celery_task_id'])

    return render(request, 'calls/_analysis_section.html', {'call': call, 'analysis': analysis})


@login_required
def ai_analysis_status(request, pk):
    """Return the current analysis status partial — used for HTMX polling."""
    from ..models import CallAnalysis

    call = get_object_or_404(Call, pk=pk)
    try:
        analysis = call.analysis
    except CallAnalysis.DoesNotExist:
        analysis = None

    return render(request, 'calls/_analysis_section.html', {'call': call, 'analysis': analysis})


@login_required
@require_POST
def cancel_ai_analysis(request, pk):
    """Cancel a pending or processing AI analysis."""
    from ..models import CallAnalysis

    call = get_object_or_404(Call, pk=pk)
    try:
        analysis = call.analysis
    except CallAnalysis.DoesNotExist:
        analysis = None

    if analysis and analysis.status in (CallAnalysis.STATUS_PENDING, CallAnalysis.STATUS_PROCESSING):
        if analysis.celery_task_id:
            try:
                from celery import current_app
                current_app.control.revoke(analysis.celery_task_id, terminate=True, signal='SIGTERM')
            except Exception:
                pass
        analysis.delete()
        analysis = None

    return render(request, 'calls/_analysis_section.html', {'call': call, 'analysis': analysis})


# ── VoIP split views ──────────────────────────────────────────────────────────

class SIPConfigForm(forms.ModelForm):
    password = forms.CharField(
        widget=forms.PasswordInput(render_value=True, attrs={"autocomplete": "new-password"}),
        required=False,
    )

    class Meta:
        model = SIPSettings
        fields = ["server_ip", "server_port", "username", "caller_id", "is_active"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance and self.instance.pk:
            self.fields["password"].initial = self.instance.password
        else:
            self.fields["password"].required = True

    def clean(self):
        cleaned = super().clean()
        password = cleaned.get("password")
        if not password and self.instance and self.instance.pk:
            password = self.instance.password

        server_ip = cleaned.get("server_ip")
        server_port = cleaned.get("server_port")
        username = cleaned.get("username")

        if not all([server_ip, server_port, username, password]):
            return cleaned
        if not cleaned.get("is_active", True):
            return cleaned

        if self.instance and self.instance.pk:
            creds_changed = (
                server_ip != self.instance.server_ip
                or server_port != self.instance.server_port
                or username != self.instance.username
                or (cleaned.get("password") and cleaned["password"] != self.instance.password)
                or cleaned.get("is_active") != self.instance.is_active
            )
            if not creds_changed:
                return cleaned

        from ..sip_validator import validate_sip_credentials
        is_valid, error = validate_sip_credentials(server_ip, server_port, username, password)
        if not is_valid:
            raise forms.ValidationError(f"SIP credential test failed: {error}")
        return cleaned

    def save(self, commit=True):
        instance = super().save(commit=False)
        password = self.cleaned_data.get("password")
        if password:
            instance.password = password
        if commit:
            instance.save()
        return instance


class SIPSoundsForm(forms.ModelForm):
    class Meta:
        model = SIPSettings
        fields = ["hold_music", "welcome_sound", "non_working_hours_sound"]

    def _validate_sound_field(self, field_name):
        sound = self.cleaned_data.get(field_name)
        if sound and hasattr(sound, "name"):
            ext = sound.name.rsplit(".", 1)[-1].lower() if "." in sound.name else ""
            if ext not in ("mp3", "wav"):
                self.add_error(field_name, "Only mp3 and wav files are supported.")

    def clean(self):
        cleaned = super().clean()
        self._validate_sound_field("hold_music")
        self._validate_sound_field("welcome_sound")
        self._validate_sound_field("non_working_hours_sound")
        return cleaned


class SIPWorkingHoursForm(forms.ModelForm):
    working_days_choices = forms.MultipleChoiceField(
        choices=WEEKDAY_CHOICES,
        widget=forms.CheckboxSelectMultiple,
        required=False,
        label="Working Days",
    )

    class Meta:
        model = SIPSettings
        fields = ["working_hours_start", "working_hours_end"]
        widgets = {
            "working_hours_start": forms.TimeInput(
                format="%H:%M", attrs={"type": "time", "class": "form-control"}
            ),
            "working_hours_end": forms.TimeInput(
                format="%H:%M", attrs={"type": "time", "class": "form-control"}
            ),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance and self.instance.pk and self.instance.working_days:
            self.fields["working_days_choices"].initial = [
                d.strip() for d in self.instance.working_days.split(",") if d.strip()
            ]

    def clean(self):
        cleaned = super().clean()
        days = cleaned.get("working_days_choices", [])
        cleaned["working_days"] = ",".join(days)
        start = cleaned.get("working_hours_start")
        end = cleaned.get("working_hours_end")
        if start and not end:
            self.add_error("working_hours_end", "End time is required when start time is set.")
        if end and not start:
            self.add_error("working_hours_start", "Start time is required when end time is set.")
        if (start and end) and not days:
            self.add_error("working_days_choices", "Select at least one working day when hours are set.")
        return cleaned

    def save(self, commit=True):
        instance = super().save(commit=False)
        instance.working_days = self.cleaned_data.get("working_days", "")
        instance.working_hours_start = self.cleaned_data.get("working_hours_start")
        instance.working_hours_end = self.cleaned_data.get("working_hours_end")
        if commit:
            instance.save()
        return instance


@login_required
@_it_admin_required
def voip_config_view(request):
    """VoIP Configuration — SIP credentials."""
    from ..asterisk_config import apply_sip_settings, apply_schedule_settings

    try:
        sip = request.user.sip_settings
    except SIPSettings.DoesNotExist:
        sip = None

    if request.method == "POST":
        if "delete" in request.POST:
            if sip:
                sip.delete()
                apply_sip_settings(None)
                apply_schedule_settings(None)
                messages.success(request, "SIP credentials deleted. Asterisk config cleared.")
            return redirect("settings:voip:voip_config")

        old_trunk = None
        if sip:
            old_trunk = {
                "server_ip": sip.server_ip,
                "server_port": sip.server_port,
                "username": sip.username,
                "is_active": sip.is_active,
                "password": sip.password,
            }

        form = SIPConfigForm(request.POST, instance=sip)
        if form.is_valid():
            trunk_changed = old_trunk is None
            if old_trunk:
                for field in ("server_ip", "server_port", "username", "is_active"):
                    if form.cleaned_data.get(field) != old_trunk[field]:
                        trunk_changed = True
                        break
                pwd = form.cleaned_data.get("password")
                if pwd and pwd != old_trunk["password"]:
                    trunk_changed = True

            obj = form.save(commit=False)
            obj.user = request.user
            obj.save()

            if trunk_changed:
                if apply_sip_settings(obj):
                    messages.success(request, "SIP credentials saved and Asterisk reloaded.")
                else:
                    messages.warning(
                        request,
                        "SIP credentials saved but Asterisk reload failed. "
                        "You may need to restart the Asterisk container.",
                    )
            else:
                messages.success(request, "Settings saved.")
            return redirect("settings:voip:voip_config")
    else:
        form = SIPConfigForm(instance=sip)

    import json
    from django.core.serializers.json import DjangoJSONEncoder
    init_data = {
        "apiUrls": {
            "sipSettings": "/settings/api/sip/",
            "voipConfig": "/settings/api/voip/config/",
            "voipSounds": "/settings/api/voip/sounds/",
            "workingHours": "/settings/api/voip/working-hours/",
        }
    }
    return render(request, "calls/voip_config.html", {
        "form": form,
        "sip_settings": sip,
        "current_section": "voip",
        "current_page": "voip_config",
        "init_data_json": json.dumps(init_data, cls=DjangoJSONEncoder),
    })


@login_required
@_it_admin_required
def voip_sounds_view(request):
    """VoIP Sounds — hold music, welcome sound, non-working hours sound."""
    from ..asterisk_config import apply_moh_settings, apply_schedule_settings

    try:
        sip = request.user.sip_settings
    except SIPSettings.DoesNotExist:
        sip = None

    if sip is None:
        messages.info(request, "Set up SIP credentials first before configuring sounds.")
        return redirect("settings:voip:voip_config")

    if request.method == "POST":
        if "remove_hold_music" in request.POST and sip.hold_music:
            sip.hold_music.delete(save=False)
            sip.hold_music = None
            sip.save()
            apply_moh_settings(sip)
            messages.success(request, "Hold music removed. Default music will be used.")
            return redirect("settings:voip:voip_sounds")

        if "remove_welcome_sound" in request.POST and sip.welcome_sound:
            sip.welcome_sound.delete(save=False)
            sip.welcome_sound = None
            sip.save()
            apply_schedule_settings(sip)
            messages.success(request, "Welcome sound removed.")
            return redirect("settings:voip:voip_sounds")

        if "remove_non_working_sound" in request.POST and sip.non_working_hours_sound:
            sip.non_working_hours_sound.delete(save=False)
            sip.non_working_hours_sound = None
            sip.save()
            apply_schedule_settings(sip)
            messages.success(request, "Non-working hours sound removed.")
            return redirect("settings:voip:voip_sounds")

        form = SIPSoundsForm(request.POST, request.FILES, instance=sip)
        if form.is_valid():
            obj = form.save()
            if "hold_music" in request.FILES:
                apply_moh_settings(obj)
            if "welcome_sound" in request.FILES or "non_working_hours_sound" in request.FILES:
                apply_schedule_settings(obj)
            messages.success(request, "Sounds saved.")
            return redirect("settings:voip:voip_sounds")
    else:
        form = SIPSoundsForm(instance=sip)

    import json
    from django.core.serializers.json import DjangoJSONEncoder
    init_data = {
        "apiUrls": {
            "sipSettings": "/settings/api/sip/",
            "voipConfig": "/settings/api/voip/config/",
            "voipSounds": "/settings/api/voip/sounds/",
            "workingHours": "/settings/api/voip/working-hours/",
        }
    }
    return render(request, "calls/voip_sounds.html", {
        "form": form,
        "sip_settings": sip,
        "current_section": "voip",
        "current_page": "voip_sounds",
        "init_data_json": json.dumps(init_data, cls=DjangoJSONEncoder),
    })


@login_required
@_it_admin_required
def voip_working_hours_view(request):
    """VoIP Working Hours — schedule and working days."""
    from ..asterisk_config import apply_schedule_settings

    try:
        sip = request.user.sip_settings
    except SIPSettings.DoesNotExist:
        sip = None

    if sip is None:
        messages.info(request, "Set up SIP credentials first before configuring working hours.")
        return redirect("settings:voip:voip_config")

    if request.method == "POST":
        form = SIPWorkingHoursForm(request.POST, instance=sip)
        if form.is_valid():
            obj = form.save()
            apply_schedule_settings(obj)
            messages.success(request, "Working hours saved.")
            return redirect("settings:voip:voip_working_hours")
    else:
        form = SIPWorkingHoursForm(instance=sip)

    import json
    from django.core.serializers.json import DjangoJSONEncoder
    init_data = {
        "apiUrls": {
            "sipSettings": "/settings/api/sip/",
            "voipConfig": "/settings/api/voip/config/",
            "voipSounds": "/settings/api/voip/sounds/",
            "workingHours": "/settings/api/voip/working-hours/",
        }
    }
    return render(request, "calls/voip_working_hours.html", {
        "form": form,
        "sip_settings": sip,
        "weekday_choices": WEEKDAY_CHOICES,
        "current_section": "voip",
        "current_page": "voip_working_hours",
        "init_data_json": json.dumps(init_data, cls=DjangoJSONEncoder),
    })
