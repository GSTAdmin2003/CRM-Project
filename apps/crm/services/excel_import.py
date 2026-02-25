"""
Excel parsing and import logic for Lead import.
"""
import re
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook

from apps.contacts.models import Company, Contact
from apps.crm.models import Lead


# Header mappings - maps header names to internal field names
# Required headers for Lead import
REQUIRED_HEADERS = ['Company ID', 'Company Name']

# Company field mappings (header -> field name)
COMPANY_HEADER_MAP = {
    'Company ID': 'legal_id',
    'Company Name': 'legal_name',
    'Brand Name': 'brand_name',
    'Company Phone': 'company_phone',
    'Company Email': 'company_email',
    'Industry': 'industry',
    'Category': 'category',
}

# Lead field mappings for Lead (header -> field name)
LEAD_HEADER_MAP = {
    'Notes': 'notes',  # Used as Lead message
}

# Contact detail fields that follow a position column
CONTACT_DETAIL_FIELDS = {'email', 'phone', 'mobile'}

MAX_ROWS = 500


def build_column_map(headers):
    """
    Build a mapping of field names to column indices based on headers.

    Detection order:
    1. Company/Lead columns are detected ONLY until we hit the first unknown column
    2. Once an unknown column is found, it starts the contact section
    3. In contact section: unknown headers = positions, Email/Phone/Mobile = contact details

    Example: Company ID | Company Name | Opportunity Title | Director | Email | Phone
    - Columns 0-2: Company/Lead data
    - Columns 3+: Contact data (Director is position, Email/Phone belong to that contact)

    Returns:
        dict: {
            'company': {field_name: col_index, ...},
            'lead': {field_name: col_index, ...},
            'contacts': {group_num: {'position': str, 'name_col': int, 'email': int, ...}, ...}
        }
    """
    column_map = {
        'company': {},
        'lead': {},
        'contacts': {},
    }

    # First pass: identify company and lead columns
    # Stop when we hit the first unknown column (start of contact section)
    contact_section_start = None

    for col_idx, header in enumerate(headers):
        if not header:
            continue

        # Strip whitespace and remove " *" suffix (used in template for required fields)
        header_clean = str(header).strip()
        if header_clean.endswith(' *'):
            header_clean = header_clean[:-2]

        # Check company headers
        if header_clean in COMPANY_HEADER_MAP:
            field_name = COMPANY_HEADER_MAP[header_clean]
            column_map['company'][field_name] = col_idx
            continue

        # Check lead headers
        if header_clean in LEAD_HEADER_MAP:
            field_name = LEAD_HEADER_MAP[header_clean]
            column_map['lead'][field_name] = col_idx
            continue

        # Unknown header - this is the start of contact section
        contact_section_start = col_idx
        break

    # Second pass: detect contact groups from contact section
    # Pattern: PositionName | Email | Phone | Mobile | NextPositionName | ...
    if contact_section_start is not None:
        contact_group = 0
        current_contact = None

        for col_idx in range(contact_section_start, len(headers)):
            header = headers[col_idx]
            if not header:
                continue

            header_clean = str(header).strip()
            if header_clean.endswith(' *'):
                header_clean = header_clean[:-2]

            header_lower = header_clean.lower()

            # Check if this is a contact detail field (email/phone/mobile)
            if header_lower in CONTACT_DETAIL_FIELDS:
                # Belongs to current contact group
                if current_contact is not None:
                    column_map['contacts'][contact_group][header_lower] = col_idx
            else:
                # This is a new contact position (e.g., "Director", "Accountant")
                contact_group += 1
                current_contact = {
                    'position': header_clean,  # The header itself is the position
                    'name_col': col_idx,       # The column value will be the contact name
                }
                column_map['contacts'][contact_group] = current_contact

    return column_map


def get_cell_value(cell):
    """Extract and clean cell value as string."""
    if cell is None or cell.value is None:
        return ''
    return str(cell.value).strip()


def get_value_by_index(row_values, col_idx):
    """Safely get value from row by column index."""
    if col_idx is None or col_idx >= len(row_values):
        return ''
    return row_values[col_idx]



def validate_row(row_data, row_number):
    """
    Validate a single row's data for Lead import.

    Returns tuple: (is_valid, errors_list)
    """
    errors = []

    company = row_data['company']
    contacts = row_data['contacts']

    # Required fields
    if not company.get('legal_id'):
        errors.append('Company ID is required')

    if not company.get('legal_name'):
        errors.append('Company Name is required')

    # Company email validation
    if company.get('company_email'):
        try:
            validate_email(company['company_email'])
        except ValidationError:
            errors.append(f'Invalid company email format: {company["company_email"]}')

    # Contact validation - if name is provided, validate email format
    for i, contact in enumerate(contacts, 1):
        if contact.get('name'):
            if contact.get('email'):
                try:
                    validate_email(contact['email'])
                except ValidationError:
                    errors.append(f'Contact {i} ({contact.get("position", "Unknown")}): Invalid email format: {contact["email"]}')

    return (len(errors) == 0, errors)


def parse_excel_file(file_obj, user):
    """
    Parse an Excel file and validate its contents for lead import.

    Args:
        file_obj: Uploaded file object
        user: The user performing the import

    Returns:
        dict: Parsed data with validation results

    Raises:
        ValueError: If file format is invalid or required headers are missing
    """
    try:
        wb = load_workbook(file_obj, data_only=True)
    except Exception as e:
        raise ValueError(f'Could not read Excel file: {str(e)}')

    ws = wb.active

    # Get headers from first row
    headers = [get_cell_value(cell) for cell in ws[1]]

    # Build column mapping from headers
    column_map = build_column_map(headers)

    # Check required headers are present
    missing_required = []
    for req_header in REQUIRED_HEADERS:
        if req_header == 'Company ID' and 'legal_id' not in column_map['company']:
            missing_required.append(req_header)
        elif req_header == 'Company Name' and 'legal_name' not in column_map['company']:
            missing_required.append(req_header)

    if missing_required:
        raise ValueError(
            f'Missing required columns: {", ".join(missing_required)}. '
            f'Please ensure your file has these column headers.'
        )

    # Count contact groups
    num_contact_groups = len(column_map['contacts'])

    # Parse data rows
    rows = []
    total_rows = 0
    valid_rows = 0
    error_rows = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Skip completely empty rows
        row_values = [get_cell_value(cell) for cell in row]
        if not any(row_values):
            continue

        total_rows += 1

        if total_rows > MAX_ROWS:
            raise ValueError(f'File exceeds maximum of {MAX_ROWS} rows. Please split into smaller files.')

        # Parse company data using column map
        company_cols = column_map['company']
        company_data = {
            'legal_id': get_value_by_index(row_values, company_cols.get('legal_id')),
            'legal_name': get_value_by_index(row_values, company_cols.get('legal_name')),
            'brand_name': get_value_by_index(row_values, company_cols.get('brand_name')),
            'company_phone': get_value_by_index(row_values, company_cols.get('company_phone')),
            'company_email': get_value_by_index(row_values, company_cols.get('company_email')),
            'industry': get_value_by_index(row_values, company_cols.get('industry')),
            'category': get_value_by_index(row_values, company_cols.get('category')),
            'exists': False,
        }

        # Check if company exists
        if company_data['legal_id']:
            company_data['exists'] = Company.objects.filter(
                legal_id=company_data['legal_id']
            ).exists()

        # Parse lead data using column map (simplified for Lead)
        lead_cols = column_map['lead']
        lead_data = {
            'notes': get_value_by_index(row_values, lead_cols.get('notes')),
        }

        # Parse contact groups using column map
        # New format: position is the header, name_col contains the contact name
        contacts = []
        for group_num in sorted(column_map['contacts'].keys()):
            contact_info = column_map['contacts'][group_num]

            # Get contact name from the position column (e.g., "Director" column contains "Jane Doe")
            contact_name = get_value_by_index(row_values, contact_info.get('name_col'))

            contact_data = {
                'name': contact_name,
                'position': contact_info.get('position', ''),  # Position is the header itself
                'email': get_value_by_index(row_values, contact_info.get('email')),
                'phone': get_value_by_index(row_values, contact_info.get('phone')),
                'mobile': get_value_by_index(row_values, contact_info.get('mobile')),
            }

            # Only add if at least name is provided
            if contact_data['name']:
                contacts.append(contact_data)

        # Create row data
        row_data = {
            'row_number': row_idx,
            'company': company_data,
            'lead': lead_data,
            'contacts': contacts,
        }

        # Validate row
        is_valid, errors = validate_row(row_data, row_idx)
        row_data['is_valid'] = is_valid
        row_data['errors'] = errors

        if is_valid:
            valid_rows += 1
        else:
            error_rows += 1

        rows.append(row_data)

    if total_rows == 0:
        raise ValueError('File contains no data rows')

    return {
        'total_rows': total_rows,
        'valid_rows': valid_rows,
        'error_rows': error_rows,
        'num_contact_groups': num_contact_groups,
        'rows': rows,
    }


def execute_import(preview_data, user):
    """
    Execute the import based on validated preview data.
    Creates Lead records (lead_type='lead', not opportunities).

    Args:
        preview_data: The parsed and validated data from parse_excel_file
        user: The user performing the import

    Returns:
        dict: Import results with statistics and per-row outcomes
    """
    results = {
        'total_processed': 0,
        'leads_created': 0,
        'companies_created': 0,
        'companies_matched': 0,
        'contacts_created': 0,
        'skipped': 0,
        'row_results': [],
    }

    for row in preview_data['rows']:
        row_result = {
            'row_number': row['row_number'],
            'company_name': row['company'].get('legal_name', ''),
            'status': 'skipped',
            'error': None,
        }

        if not row['is_valid']:
            row_result['status'] = 'skipped'
            row_result['error'] = '; '.join(row['errors'])
            results['skipped'] += 1
            results['row_results'].append(row_result)
            continue

        results['total_processed'] += 1

        try:
            with transaction.atomic():
                company_data = row['company']
                lead_data = row['lead']
                contacts_data = row['contacts']

                # 1. Get or create Company
                try:
                    company = Company.objects.get(legal_id=company_data['legal_id'])
                    results['companies_matched'] += 1

                    # Update empty fields with provided data
                    updated = False
                    if company_data.get('brand_name') and not company.brand_name:
                        company.brand_name = company_data['brand_name']
                        updated = True
                    if company_data.get('company_phone') and not company.company_phone:
                        company.company_phone = company_data['company_phone']
                        updated = True
                    if company_data.get('company_email') and not company.company_email:
                        company.company_email = company_data['company_email']
                        updated = True
                    if company_data.get('industry') and not company.industry:
                        company.industry = company_data['industry']
                        updated = True
                    if company_data.get('category') and not company.category:
                        company.category = company_data['category']
                        updated = True

                    if updated:
                        company.updated_by = user
                        company.save()

                except Company.DoesNotExist:
                    company = Company.objects.create(
                        legal_id=company_data['legal_id'],
                        legal_name=company_data['legal_name'],
                        brand_name=company_data.get('brand_name', ''),
                        company_phone=company_data.get('company_phone', ''),
                        company_email=company_data.get('company_email', ''),
                        industry=company_data.get('industry', ''),
                        category=company_data.get('category', ''),
                        created_by=user,
                        updated_by=user,
                    )
                    results['companies_created'] += 1

                # 2. Create Contacts
                lead_contact = None
                for contact_data in contacts_data:
                    if not contact_data.get('name'):
                        continue

                    contact = Contact.objects.create(
                        company=company,
                        name=contact_data['name'],
                        position=contact_data.get('position', ''),
                        email=contact_data.get('email', ''),
                        phone=contact_data.get('phone', ''),
                        mobile=contact_data.get('mobile', ''),
                    )
                    results['contacts_created'] += 1

                    if lead_contact is None:
                        lead_contact = contact

                # 3. Create Lead (lead_type='lead')
                message = lead_data.get('notes', '') or ''

                Lead.objects.create(
                    lead_type=Lead.TYPE_LEAD,
                    company=company,
                    contact=lead_contact,
                    message=message,
                    status='new',
                    assigned_to=user,
                    sales_team=getattr(user, 'sales_team', None),
                    created_by=user,
                )
                results['leads_created'] += 1

                row_result['status'] = 'created'

        except Exception as e:
            row_result['status'] = 'error'
            row_result['error'] = str(e)

        results['row_results'].append(row_result)

    return results
